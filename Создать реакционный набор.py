from tkinter import *
from tkinter.ttk import Combobox
import openpyxl
from DATA.components.read_components import read_components_from_xlsx

fold = 'DATA/reactions'


class New_Reactions_Window(Toplevel):
    def __init__(self, all_components: dict):
        super().__init__()
        self.book = openpyxl.Workbook()

        self.row = 2
        self.parameter = True
        while self.parameter:
            big_frame = Frame(self)
            frame = Frame(big_frame)
            Label(frame, text='').grid(column=0, row=1)
            Label(frame, text='Стехиометрия').grid(column=0, row=2)
            Label(frame, text='Порядок реакции по компоненту').grid(column=0, row=3)
            frame.grid(column=0, row=1)

            self.dict_comp = {}
            Label(big_frame, text='Компоненты').grid(column=1, row=0)
            Label(big_frame, text='A').grid(column=2, row=0)
            self.entry_A = Entry(big_frame, width=15, borderwidth=2)
            self.entry_A.grid(column=2, row=1)
            self.entry_E = Entry(big_frame, width=15, borderwidth=2)
            self.entry_E.grid(column=3, row=1)
            Label(big_frame, text='E').grid(column=3, row=0)


            frame = Frame(big_frame)
            for i in range(4):
                self.dict_comp[i] = {'list': Combobox(frame, values=tuple(all_components.keys()), width=12),
                                     'value': Entry(frame, width=15, borderwidth=2),
                                     'n': Entry(frame, width=15, borderwidth=2)}
                self.dict_comp[i]['list'].grid(column=i, row=1)
                self.dict_comp[i]['value'].grid(column=i, row=2)
                self.dict_comp[i]['n'].grid(column=i, row=3)
            frame.grid(column=1, row=1)

            big_frame.pack()
            Button(self, text='Готово', command=self.done).pack()
            Button(self, text='Добавить еще реакцию', command=self.add).pack()
            self.wait_window()
            self.row += 4

    def done(self):
        self.parameter = False
        self.save_reaction()
        self.destroy()

    def add(self):
        self.parameter = True
        self.save_reaction()
        self.destroy()
        super().__init__()

    def save_reaction(self):
        sheet = self.book.active
        for i in self.dict_comp.keys():
            if self.dict_comp[i]['list'].get() != '':
                sheet.cell(column=i + 2, row=self.row - 1).value = self.dict_comp[i]['list'].get()
                sheet.cell(column=i + 2, row=self.row).value = int(self.dict_comp[i]['value'].get())
                try:
                    sheet.cell(column=i + 2, row=self.row + 1).value = int(self.dict_comp[i]['n'].get())
                except ValueError:
                    sheet.cell(column=i + 2, row=self.row + 1).value = 0
        sheet.cell(column=6, row=self.row - 1).value = 'A'
        sheet.cell(column=6, row=self.row + 1).value = float(self.entry_A.get())
        sheet.cell(column=7, row=self.row - 1).value = 'E'
        sheet.cell(column=7, row=self.row + 1).value = float(self.entry_E.get())
        self.book.save(f'{fold}/Name.xlsx')


if __name__ == '__main__':
    all = read_components_from_xlsx()
    ra = New_Reactions_Window(all)
