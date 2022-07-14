import openpyxl
from src.Reactions import Reactions


def read_reactions_from_xlsx(folder='./DATA/reactions', filename='Mol_Reactions.xlsx'):
    xlsx = openpyxl.load_workbook(f'{folder}/{filename}', data_only=True)
    sheet = xlsx.active

    reactions = Reactions(filename)

    parameters = {}

    row_id = 1
    while True:
        id = sheet.cell(row=row_id, column=1).value
        if id is None:
            break

        parameters['components_reaction'] = [col for col in sheet.iter_cols(min_row=row_id, max_row=row_id + 2,
                                                                            min_col=2, values_only=True)
                                             if col[0] is not None]
        parameters['A'] = sheet.cell(row=row_id + 3, column=2).value
        parameters['E'] = sheet.cell(row=row_id + 4, column=2).value

        reactions.add_reaction(id, parameters)

        row_id += 6

    return reactions
