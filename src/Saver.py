import openpyxl


class Saver:
    """Класс предназначен для сохранения введенных данных и результатов в таблицу Excel"""

    ROW_COMP = 1
    COL_COMP = 1

    _ROW = 1
    _NEXT_ROW = 5
    _INIT_COMPONENTS_COLUMN = 1
    _REACTOR_COLUMN = 8
    _RESULT_COMPONENTS_COLUMN = 14

    _COL_TIME = 2

    def __init__(self, model):
        self.model = model
        reaction_filename = self.model.get_reactions().filename.split('.')[0]
        reactor_filename = self.model.get_cascade().filename.split('.')[0]
        self.filename = f'{reaction_filename} in {reactor_filename}.xlsx'

        # Создание Excel файла
        self.xlsx = openpyxl.Workbook()
        self.sheet_reactor = self.xlsx.active
        self.sheet_reactor.title = 'Реактор'
        # self.sheet_reactions = self.xlsx.create_sheet(title='Реакции')
        # self.sheet_short_result = self.xlsx.create_sheet(title='Результат')
        self.sheet_result = self.xlsx.create_sheet(title='Сводка')
        self.sheet_time = self.xlsx.create_sheet(title='Время')
        self.set_composition_on_sheet(self.sheet_result)
        self.set_composition_on_sheet(self.sheet_time)

    def set_composition_on_sheet(self, sheet):
        row = 1
        sheet.cell(row=row, column=1, value='Компонент')
        sheet.cell(row=row, column=2, value='Молярная масса,\nкг/кмоль')
        for component in self.model.get_flow().get_composition():
            row += 1
            sheet.cell(row=row, column=1, value=component.name)
            sheet.cell(row=row, column=2, value=component.molar_mass).number_format = '0.0000'

    def reactor(self, reactor):
        VALUES_REACTOR = [
            ['Параметр', reactor.name],
            ['Температура, С', reactor.temperature_string],
            ['Давление, кПа', reactor.pressure_string],
            ['Объем реактора, м3', reactor.volume],
            ['Кол-во секций, шт', reactor.sections_count],
            ['Время пребывания, с', reactor.resident_time]
        ]
        ROUNDINGS = ['', '0.00', '0.00', '0.00', '0', '0.000000']

        row, column = self._ROW, self._REACTOR_COLUMN
        for current_row, parameters in enumerate(VALUES_REACTOR):
            for current_column, parameter in enumerate(parameters):
                self.sheet_reactor.cell(row=row + current_row, column=column + current_column, value=parameter).number_format = ROUNDINGS[current_row]
        self._ROW += 7
        self.xlsx.save(f'{self.filename}')

    @staticmethod
    def write_in_xlsx(data, start_row, start_col, sheet) -> None:
        row = start_row
        for row_list in data:
            col = start_col
            for item in row_list:
                sheet.cell(row=row, column=col, value=item).number_format = '0.0000'
                col += 1
            row += 1

    def save(self):
        self.xlsx.save(f'{self.filename}')
